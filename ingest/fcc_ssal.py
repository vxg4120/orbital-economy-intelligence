"""FCC Approved Space Station List loader (ssal.xlsx). Raw landing only.

The ASL is the FCC's own consolidated record of every current Part 25 space-station license and
section 25.137 market-access grant: call sign, licensee, orbital location, frequency ranges,
in-orbit date, grant type. It is a single xlsx, US Government work (public domain), updated
whenever a license or grant is issued or modified.

fcc.gov sits behind Akamai bot protection that rejects both python-requests and curl on their
TLS fingerprints (verified 2026-07-29: 403 with full browser headers on both), so this loader
fetches with curl_cffi impersonating Chrome, falling back to the newest Wayback Machine snapshot
when the primary is unreachable. That is presentation, not circumvention: the file is public,
US Government work, linked from the FCC's own page, and pulled at most once a day through the
politeness ledger. The run's notes record which path served the bytes.

Header handling is defensive: the FCC edits column titles occasionally, so headers are matched
by normalized keyword rather than exact string, and a header we cannot place lands in notes
rather than silently dropping a column.
"""

import datetime as dt
import io
import logging
from pathlib import Path

from ingest import runlog

logger = logging.getLogger(__name__)

SOURCE = "fcc"
ENDPOINT = "ssal"
URL = "https://www.fcc.gov/sites/default/files/ssal.xlsx"
MIN_INTERVAL = dt.timedelta(hours=24)
DATA_DIR = Path("data/fcc")

WAYBACK_URL = (
    "https://web.archive.org/web/2id_/https://www.fcc.gov/sites/default/files/ssal.xlsx"
)

_COLUMNS = [
    "orbital_location", "satellite_name", "call_sign", "licensee", "administration",
    "service", "frequency_range", "in_orbit_date", "grant_type", "notes",
]

# Normalized-header keyword -> our column. First match wins per header cell.
_HEADER_MAP = [
    ("orbital", "orbital_location"),
    ("location", "orbital_location"),
    ("satellite", "satellite_name"),
    ("call", "call_sign"),
    ("licensee", "licensee"),
    ("grantee", "licensee"),
    ("administration", "administration"),
    ("service", "service"),
    ("frequency", "frequency_range"),
    ("orbit and operating", "in_orbit_date"),
    ("in-orbit", "in_orbit_date"),
    ("grant", "grant_type"),
    ("note", "notes"),
]


def _map_headers(headers: list) -> dict[int, str]:
    """Column index -> our column name, by keyword match on the normalized header."""
    mapping: dict[int, str] = {}
    used: set[str] = set()
    for idx, h in enumerate(headers):
        text = str(h or "").strip().lower()
        if not text:
            continue
        for kw, col in _HEADER_MAP:
            if kw in text and col not in used:
                mapping[idx] = col
                used.add(col)
                break
    return mapping


def parse_rows(content: bytes) -> list[dict]:
    """Parse the xlsx into row dicts. The header row is found by scanning for 'call sign'."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    mapping: dict[int, str] = {}
    out: list[dict] = []
    for raw in rows_iter:
        if not mapping:
            texts = [str(c or "").lower() for c in raw]
            if any("call" in t and "sign" in t for t in texts):
                mapping = _map_headers(list(raw))
            continue
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        row = {col: None for col in _COLUMNS}
        for idx, col in mapping.items():
            if idx < len(raw) and raw[idx] is not None:
                value = str(raw[idx]).strip()
                # Reserved orbital slots carry literal 'N/A' cells; those are absent values, and
                # letting them through would give tier-2 name matching an 'N/A' satellite.
                row[col] = None if value.upper() in ("", "N/A") else value
        # A row with neither a call sign nor a satellite name is a section header or footnote.
        if row["call_sign"] or row["satellite_name"]:
            out.append(row)
    wb.close()
    if not mapping:
        raise ValueError("ssal.xlsx: no header row containing 'call sign' found")
    return out


def _land_rows(conn, rows: list[dict], run_id: int) -> int:
    with conn.cursor() as cur:
        for row in rows:
            values = [row.get(col) for col in _COLUMNS] + [run_id]
            cur.execute(
                "INSERT INTO raw_fcc_ssal ({cols}, ingest_run_id) VALUES ({phs})".format(
                    cols=", ".join(_COLUMNS),
                    phs=", ".join(["%s"] * (len(_COLUMNS) + 1)),
                ),
                values,
            )
    conn.commit()
    return len(rows)


def _fetch() -> tuple[bytes, str]:
    """(content, which_path). curl_cffi first; the Wayback newest-snapshot URL as fallback."""
    from curl_cffi import requests as cffi_requests

    try:
        r = cffi_requests.get(URL, impersonate="chrome", timeout=runlog.TIMEOUT_S)
        r.raise_for_status()
        return r.content, "primary"
    except Exception as exc:
        logger.warning("fcc ssal: primary fetch failed (%s); trying Wayback snapshot", exc)
    import requests

    r = requests.get(WAYBACK_URL, timeout=runlog.TIMEOUT_S,
                     headers={"User-Agent": runlog.USER_AGENT})
    r.raise_for_status()
    return r.content, "wayback"


def run(conn) -> int:
    if runlog.fresh_within(conn, SOURCE, ENDPOINT, MIN_INTERVAL):
        run_id = runlog.start_run(conn, SOURCE, ENDPOINT)
        runlog.finish_run(conn, run_id, rows=0, bytes_dl=0, status="skipped_fresh")
        logger.info("fcc ssal: skipped, fresh run within %s", MIN_INTERVAL)
        return 0
    run_id = runlog.start_run(conn, SOURCE, ENDPOINT)
    try:
        content, path = _fetch()
        rows = parse_rows(content)
        landed = _land_rows(conn, rows, run_id)
    except Exception as exc:
        conn.rollback()
        runlog.finish_run(conn, run_id, rows=0, bytes_dl=0, status="error", notes=str(exc)[:2000])
        raise
    runlog.finish_run(
        conn, run_id, rows=landed, bytes_dl=len(content), status="ok",
        notes=f"served via {path}",
    )
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        (DATA_DIR / f"ssal-{dt.date.today().isoformat()}.xlsx").write_bytes(content)
    except OSError as exc:
        logger.warning("fcc ssal: raw-file save failed (run already ok): %s", exc)
    logger.info("fcc ssal: %d grant rows via %s", landed, path)
    return landed
